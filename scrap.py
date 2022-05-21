import openpyxl as px 
import jeraconv.jeraconv
import re 
import sys
import os 
from selenium import webdriver
from selenium.common.exceptions import ElementNotInteractableException, NoSuchElementException, WebDriverException
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup as bs
import time

#TODO:[hotfix/RENEW20220521]ページ上の要素のセレクタ変更に伴う更新

class Scraping:
    
    def __init__(self, path):
        #create new book
        self.path = path
        if os.path.isfile(self.path):
            self.book = px.load_workbook(path)
            self.sheet = self.book.worksheets[0]
        else:
            self.book = px.Workbook()
            self.sheet = self.book.worksheets[0]
            self.ready_book()
        #initialization
        self.options = webdriver.ChromeOptions()
        self.options.add_argument("start-maximized")
        self.options.add_argument("enable-automation")
        #self.options.add_argument("--headless")
        self.options.add_argument('--lang=ja-JP')
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--proxy-server='direct://'")
        self.options.add_argument("--proxy-bypass-list=*")
        self.options.add_argument("--disable-infobars")
        self.options.add_argument('--disable-extensions')
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_argument("--disable-browser-side-navigation")
        self.options.add_argument("--disable-gpu")
        self.options.add_argument('--ignore-certificate-errors')
        self.options.add_argument('--ignore-ssl-errors')
        prefs = {"profile.default_content_setting_values.notifications": 2}
        self.options.add_experimental_option("prefs", prefs)
        browser_path = resource_path('chrome-win/chrome.exe')
        self.options.binary_location = browser_path
        self.resultcnt = 1
        self.driver_path = resource_path('chromedriver_win32/chromedriver.exe')
        #self.driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        #self.driver.set_window_size('1200', '1000')
        self.count = 0
        self.end_flg = False

    def search(self, area, honten):
        def select_choice(select_text, element_id):
            choice = self.driver.find_element_by_id(element_id)
            select = Select(choice)
            return select.select_by_visible_text(select_text)
        
        self.driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        self.driver.set_window_size('1200', '1000')
        wait = WebDriverWait(self.driver, 20)
        self.area = area
        self.honten = honten
        self.driver.get('https://etsuran.mlit.go.jp/TAKKEN/kensetuKensaku.do')
        wait.until(EC.visibility_of_all_elements_located)
        try:#接続エラー発生時：初回検索操作時およびrestart()➞ページ表示時
            sys_down = self.driver.find_element_by_id('Red')#接続エラーの要素をキャッチ
            self.restart()
        except NoSuchElementException:#要素をキャッチ出来なければなにもせず通常処理へ
            if honten:
                select_choice('本店', 'choice') 
            else:
                pass
            select_choice(area, 'kenCode')
            select_choice('50', 'dispCount')
            search_btn = self.driver.find_element_by_css_selector('#input > div:nth-child(6) > div:nth-child(5)')
            search_btn.click()
        
    
    def restart(self):
        self.book.save(self.path)
        self.driver.delete_all_cookies()
        self.driver.quit()
        time.sleep(10)
        self.search(self.area, self.honten)

    def scrap(self):
        wait = WebDriverWait(self.driver, 20)
        wait.until(EC.visibility_of_all_elements_located)
        self.resultcnt = self.driver.find_element_by_css_selector('#container_cont > div.result.clr > p').text
        self.resultcnt = self.resultcnt.replace("検索結果：", "")
        self.resultcnt = self.resultcnt.replace("件", "")
        self.resultcnt = re.sub("\n\d+目～\d+目までを表示", "", self.resultcnt)
        print(self.resultcnt)
        self.resultcnt = int(self.resultcnt)
        self.count = 1
        res_count = self.driver.find_element_by_id('pageListNo1')
        select = Select(res_count)
        all_options = select.options
        loop_count = len(all_options)
        index = self.sheet.max_row + 1
        for i in range(1, loop_count):
            try:#接続エラー発生時：restart->現在のi番目のリストからやり直し。ページ遷移時
                    sys_down = self.driver.find_element_by_id('Red')
                    self.restart()
                    menu = self.driver.find_element_by_css_selector('#pageListNo1')
                    select = Select(menu)
                    select.select_by_value(str(i))
            except NoSuchElementException:
                pass
            if i % 10 == 0:#10ページに1回ブラウザを再起動しメモリをクリア。
                self.restart()
                menu = self.driver.find_element_by_css_selector('#pageListNo1')
                select = Select(menu)
                select.select_by_value(str(i))
            for j in range(2, 52):
                try:#接続エラー発生時：restart->現在のi番目のリストからやり直し。会社選択時
                    sys_down = self.driver.find_element_by_id('Red')
                    self.restart()
                    menu = self.driver.find_element_by_css_selector('#pageListNo1')
                    select = Select(menu)
                    select.select_by_value(str(i))
                except NoSuchElementException:
                    pass
                
                #InfoScrap here
                try:
                    wait.until(EC.visibility_of_all_elements_located)
                    company = self.driver.find_element_by_css_selector('#container_cont > table > tbody > tr:nth-child(' + str(j) +  ') > td:nth-child(4) > a')
                    #ElementClickInterceptException対策(画面外要素クリック対策)
                    self.driver.execute_script('arguments[0].click();', company)
                    #company.click()
                except (ElementNotInteractableException, NoSuchElementException):
                    pass
                else:
                    try:#接続エラー発生時：restart->現在のi番目のリストからやり直し。選択➞情報抽出時
                        sys_down = self.driver.find_element_by_id('Red')
                        self.restart()
                        menu = self.driver.find_element_by_css_selector('#pageListNo1')
                        select = Select(menu)
                        select.select_by_value(str(i))
                        wait.until(EC.visibility_of_all_elements_located)
                        company = self.driver.find_element_by_css_selector('#container_cont > table > tbody > tr:nth-child(' + str(j) +  ') > td:nth-child(4) > a')
                        #ElementClickInterceptException対策(画面外要素クリック対策)
                        self.driver.execute_script('arguments[0].click();', company)
                    except NoSuchElementException:
                        pass
                    wait.until(EC.visibility_of_all_elements_located)
                    html = self.driver.page_source
                    try:
                        self.__extraction(html, index)
                        print(self.sheet.max_row)
                    except:#抽出処理時のみ全例外をスルー
                        pass
                    self.driver.back()                        
                index += 1
                self.count += 1
            #MaxRetryError                                   
            #wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#container_cont > div.result.clr > div:nth-child(5) > img")))
            wait.until(EC.visibility_of_all_elements_located)
            next_btn = self.driver.find_element_by_css_selector('#container_cont > div.result.clr > div:nth-child(5) > img')
            next_btn.click()
        self.book.save(self.path)
        print("saved")
        print(self.sheet.max_row)
        self.driver.quit()
       

    def ready_book(self):
        col_list = [
            "許可年月", 
            "許可番号簡易",
            "商号又は名称（カナ）",
            "商号又は名称（詳細）",	
            "代表者の氏名（カナ）",	
            "代表者の氏名（漢字）",
            "郵便番号",
            "都道府県コード",
            "都道府県",
            "市区町村・番地・建物",	
            "電話番号",
            "法人・個人区分",
            "資本金額",	
            "建設業以外の兼業の有無",	
            "土木工事業",
            "建築工事業",
            "大工工事業",
            "左官工事業",
            "とび・土工工事業",
            "石工事業",
            "屋根工事業",
            "電気工事業",
            "管工事業",	
            "タイル・れんが・ブロツク工事業",
            "鋼構造物工事業", 
            "鉄筋工事業",
            "舗装工事業",
            "しゆんせつ工事業",
            "板金工事業",
            "ガラス工事業",
            "塗装工事業",
            "防水工事業",
            "内装仕上工事業",
            "機械器具設置工事業",
            "熱絶縁工事業",
            "電気通信工事業",
            "造園工事業",
            "さく井工事業",
            "建具工事業",
            "水道施設工事業",
            "消防施設工事業",
            "清掃施設工事業",
            "許可の有効期間",
        ]
        for c in range(1, len(col_list)+1):
            self.sheet.cell(row=1, column=c, value=col_list[c-1])
        self.sheet.freeze_panes = "A2"
        self.book.save(self.path)

    def __extraction(self, html, index):
        soup = bs(html, 'lxml')
        #print(html)
        perm_day_tag = soup.select_one("div.clr > div > div.scroll-pane > table.re_summ_4 > tbody > tr > td > a")
        perm_day = perm_day_tag.get_text() if perm_day_tag != None else None
        perm_day = self.wareki_conv(perm_day) if perm_day != None else None
        self.sheet.cell(row=index, column=1, value=perm_day)
        #print(index, end=" ")
        #print(self.sheet.cell(row=index, column=1).value)
        perm_num_str_tag = soup.select_one("#input > div.clr > table > tbody > tr > td")
        perm_num_str = perm_num_str_tag.get_text() if perm_num_str_tag != None else None
        perm_num = perm_num_str.split("　")[1] if perm_num_str != None else None
        self.sheet.cell(row=index, column=2, value=perm_num)
        
        name_kana_tag = soup.select_one("#input > div:nth-child(1) > table > tbody > tr:nth-child(2) > td > p")
        name_kana = name_kana_tag.get_text() if name_kana_tag != None else None
        self.sheet.cell(row=index, column=3, value=name_kana)
        
        com_name_tag = soup.select_one(("#input > div:nth-child(1) > table > tbody > tr:nth-child(2) > td"))
        com_name_str = com_name_tag.get_text() if com_name_tag != None else None
        com_name = com_name_str.replace(name_kana, "") if com_name_str != None else None
        self.sheet.cell(row=index, column=4, value=com_name)

        ceo_kana_tag = soup.select_one("#input > div:nth-child(1) > table > tbody > tr:nth-child(3) > td > p")
        ceo_kana = ceo_kana_tag.get_text() if ceo_kana_tag != None else None
        self.sheet.cell(row=index, column=5, value=ceo_kana)

        ceo_name_tag = soup.select_one("#input > div:nth-child(1) > table > tbody > tr:nth-child(3) > td")
        ceo_name_str = ceo_name_tag.get_text() if ceo_name_tag != None else None
        ceo_name = ceo_name_str.replace(ceo_kana, "") if ceo_name_str != None else None
        self.sheet.cell(row=index, column=6, value=ceo_name)
        #住所処理系
        address_data_tag = soup.select_one('#input > div:nth-child(1) > table > tbody > tr:nth-child(4) > td')
        address_data = address_data_tag.get_text()
        post_obj = re.search('[0-9]{3}-[0-9]{4}', address_data)
        post_num = post_obj.group()
        self.sheet.cell(row=index, column=7, value=post_num)
        comp_address_data = re.split('〒[0-9]{3}-[0-9]{4}', address_data)#住所のみの文字列へ変換
        pref_obj = re.search('東京都|北海道|(?:京都|大阪)府|.{2,3}県', comp_address_data[1])
        pref = pref_obj.group()
        self.sheet.cell(row=index, column=8, value=self.call_jis_code(pref))
        self.sheet.cell(row=index, column=9, value=pref)
        muni = re.split('東京都|北海道|(?:京都|大阪)府|.{2,3}県', comp_address_data[1])#市区町村
        self.sheet.cell(row=index, column=10, value=muni[1])
        
        tel_tag = soup.select_one('#input > div:nth-child(1) > table > tbody > tr:nth-child(5) > td')
        tel = tel_tag.get_text() if tel_tag != None else None
        self.sheet.cell(row=index, column=11, value=tel)

        com_class_tag = soup.select_one('table.re_summ_2 > tbody > tr:nth-child(1) > td')
        com_class = com_class_tag.get_text() if com_class_tag != None else None
        self.sheet.cell(row=index, column=12, value=com_class)

        com_money_tag = soup.select_one('table.re_summ_2 > tbody > tr.tdnum > td')
        com_money = com_money_tag.get_text() if com_money_tag != None else None
        if com_money != None:
            com_money = com_money.replace(",", "")
            com_money = com_money.replace("千円", "")
            com_money = int(com_money)
        self.sheet.cell(row=index, column=13, value=com_money)

        kengyou_tag = soup.select_one('table.re_summ_2 > tbody > tr:nth-child(3) > td')
        kengyou = kengyou_tag.get_text() if kengyou_tag != None else None
        self.sheet.cell(row=index, column=14, value=kengyou)

        #表処理系
        perm_data = soup.select('#input > table:nth-child(6) > tbody > tr.re_summ_odd > td')
        for c in range(15, 43):
            i = c - 15
            perm_class = perm_data[i].get_text()
            if perm_class in ("1", "2") :
                self.sheet.cell(row=index, column=c, value="●")
        
        perm_period = soup.select_one('div.clr > div > table.re_summ_5 > tbody > tr > td').get_text()
        self.sheet.cell(row=index, column=self.sheet.max_column, value=perm_period)

    def wareki_conv(self, day_data:str):
        j2w = jeraconv.jeraconv.J2W()
        day_data = day_data.replace("R", "令和")
        day_data = day_data.replace("H", "平成")
        day_array = day_data.split("/")
        day_array[0] += "年"
        year = j2w.convert(day_array[0])
        day = str(year) + "/" + day_array[1] + "/" + day_array[2]
        return day

    def call_jis_code(self, key):
        pref_jiscode = {
            "北海道": 1,
            "青森県": 2,
            "岩手県": 3,
            "宮城県": 4,
            "秋田県": 5,
            "山形県": 6,
            "福島県": 7,
            "茨城県": 8,
            "栃木県": 9,
            "群馬県": 10,
            "埼玉県": 11,
            "千葉県": 12,
            "東京都": 13,
            "神奈川県": 14,
            "新潟県": 15,
            "富山県": 16,
            "石川県": 17,
            "福井県": 18,
            "山梨県": 19,
            "長野県": 20,
            "岐阜県": 21,
            "静岡県": 22,
            "愛知県": 23,
            "三重県": 24,
            "滋賀県": 25,
            "京都府": 26,
            "大阪府": 27,
            "兵庫県": 28,
            "奈良県": 29,
            "和歌山県": 30,
            "鳥取県": 31,
            "島根県": 32,
            "岡山県": 33,
            "広島県": 34,
            "山口県": 35,
            "徳島県": 36,
            "香川県": 37,
            "愛媛県": 38,
            "高知県": 39,
            "福岡県": 40,
            "佐賀県": 41,
            "長崎県": 42,
            "熊本県": 43,
            "大分県": 44,
            "宮崎県": 45,
            "鹿児島県": 46,
            "沖縄県": 47
        }
        code = pref_jiscode[key]
        return code

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)



if __name__ == "__main__":

    scrap = Scraping('./test.xlsx')
    scrap.search('14 神奈川県',True)
    scrap.scrap()
    """
    main("datas/Hokaido.xlsx", "01 北海道")
    main("daats/Tokyo.xlsx", "13 東京都")
    main("datas/Tiba.xlsx", "12 千葉県")
    main("datas/Saitama.xlsx", "11 埼玉県")
    main("datas/Osaka.xlsx", "27 大阪府")
    main("datas/Kyoto.xlsx", "26 京都府")
    main("datas/Hyogo.xlsx", "28 兵庫県")
    main("datas/Fucuoka.xlsx", "40 福岡県")
    main("datas/Saga.xlsx", "41 佐賀県")
    """

    """
    task = [
        ("./Kanagawa.xlsx", "14 神奈川県"),
        ("./Hokaido.xlsx", "01 北海道"),
        ("./Tokyo.xlsx", "13 東京都"),
        ("./Tiba.xlsx", "12 千葉県"),
        ("./Saitama.xlsx", "11 埼玉県"),
        ("./Osaka.xlsx", "27 大阪府"),
        ("./Kyoto.xlsx", "26 京都府"),
        ("./Hyogo.xlsx", "28 兵庫県"),
        ("./Fucuoka.xlsx", "40 福岡県"),
        ("./Saga.xlsx", "41 佐賀県"),
    ]
    ths = []
    for tk in task:
        th = threading.Thread(target=main, args=tk)
        ths.append(th)
    for th in ths:
        th.start()
    """
        




